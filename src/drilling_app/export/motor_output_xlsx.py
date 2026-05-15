"""Export motor output data to the legacy Motor Output.xlsx wide-column format."""
from __future__ import annotations

from io import BytesIO
from typing import Sequence

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from ..models import BhaRun, MotorOutput


def build_export(bha_runs: Sequence[BhaRun]) -> bytes:
    """Generate legacy Motor Output.xlsx with 8 columns per BHA run."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Motor Output"

    col = 1
    for run in bha_runs:
        well = run.well
        well_label = f"{well.name}"
        if run.bha_name:
            well_label += f"_{run.bha_name}"
        if run.motor_bent_deg:
            well_label += f"_{run.motor_bent_deg}degBent"
        if run.hole_size_in:
            well_label += f"_{run.hole_size_in}in"

        # Row 1: well name
        ws.cell(row=1, column=col, value=well_label).font = Font(bold=True)
        # Row 2: column headers
        headers = ["Svy MD", "Incl", "Azmth", "DLS", "Slide footage", "Motor Output", "Stand", ""]
        for j, h in enumerate(headers):
            ws.cell(row=2, column=col + j, value=h)
        # Row 3: units
        stand = run.stand_length_ft
        units = ["ft", "deg", "deg", "deg/100ft", "ft", "deg/ft", str(int(stand)), ""]
        for j, u in enumerate(units):
            ws.cell(row=3, column=col + j, value=u)

        # Data rows
        for i, mo in enumerate(run.motor_outputs, start=4):
            ws.cell(row=i, column=col, value=mo.svy_md_ft)
            ws.cell(row=i, column=col + 1, value=mo.incl_deg)
            ws.cell(row=i, column=col + 2, value=mo.azmth_deg)
            ws.cell(row=i, column=col + 3, value=mo.dls_deg_per_100ft)
            ws.cell(row=i, column=col + 4, value=mo.slide_footage_ft)
            ws.cell(row=i, column=col + 5, value=mo.motor_output_deg_per_ft)
            ws.cell(row=i, column=col + 6, value=mo.full_stand_deg)

        col += 8

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
