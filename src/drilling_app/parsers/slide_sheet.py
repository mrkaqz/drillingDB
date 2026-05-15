"""Parser for DOX slide sheet / SLB Steering Sheet Excel files.

Handles multiple format variants:
- Modern DOX (ft, deg/100ft, PPG) with Operation Mode in any column
- Older SLB Steering Sheets (m, deg/10m, SG, uppercase modes, abbreviated labels)
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Optional

import openpyxl
from dateutil import parser as dateutil_parser

from .label_map import lookup_header, lookup_ops_col, normalize
from ..compute.units import (
    depth_to_ft, dls_to_100ft, mw_to_ppg,
    detect_depth_unit, detect_dls_unit, detect_mw_unit,
)

RECOGNISED_MODES = {"Sliding", "Rotating", "Circulating", "Tripping", "Reaming", "Backreaming"}


@dataclass
class SurveyRow:
    sequence: int
    svy_md_ft: float
    incl_deg: float
    azmth_deg: float
    dls_deg_per_100ft: float


@dataclass
class IntervalRow:
    sequence: int
    md_from_ft: float
    md_to_ft: float
    mode: str
    course_ft: Optional[float] = None
    calc_rop: Optional[float] = None
    tf_mode: Optional[str] = None
    tf_angle: Optional[float] = None
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None


@dataclass
class ParsedSlideSheet:
    header: dict[str, Any]
    surveys: list[SurveyRow]
    intervals: list[IntervalRow]
    source_unit_system: str  # "field" | "metric"
    source_units: dict[str, str]
    source_filename: str = ""


def _cell_str(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    if hasattr(v, "text"):  # ArrayFormula or similar
        return ""
    return str(v).strip()


def _parse_date(v) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        return dateutil_parser.parse(s, dayfirst=False)
    except Exception:
        return None


def _safe_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _normalize_mode(raw: str) -> str:
    normalised = raw.strip().title()
    return normalised if normalised in RECOGNISED_MODES else normalised


def _scan_header_block(ws, row_start: int, row_end: int) -> dict[str, Any]:
    """Walk rows 4..14 and collect label→value pairs using the synonym table.

    Looks ahead up to 2 cells for the value — handles layouts where a blank
    column separates the label from its value (e.g. GYR-format slide sheets).
    """
    result: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=row_start, max_row=row_end):
        for i, cell in enumerate(row):
            raw = _cell_str(cell)
            if not raw:
                continue
            key = lookup_header(raw)
            if key:
                for offset in (1, 2):
                    if i + offset < len(row):
                        neighbour = row[i + offset]
                        if neighbour.value is not None and _cell_str(neighbour) != "":
                            result[key] = neighbour.value
                            result[f"_label_{key}"] = raw
                            break
    return result


def _find_ops_header_row(ws) -> Optional[int]:
    """Find the row containing 'Start Time' or 'MD From'."""
    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            raw = _cell_str(cell).lower()
            if raw in ("start time", "md from", "md to"):
                return cell.row
    return None


def _build_col_map(ws, header_row: int) -> dict[str, int]:
    """Map canonical column names to 1-based column indices."""
    col_map: dict[str, int] = {}
    for cell in ws[header_row]:
        raw = _cell_str(cell)
        if raw:
            key = lookup_ops_col(raw)
            if key and key not in col_map:
                col_map[key] = cell.column
    return col_map


def _get_units_row(ws, header_row: int, col_map: dict[str, int]) -> dict[str, str]:
    """Read the row directly below the header row to extract unit strings."""
    units: dict[str, str] = {}
    units_row = header_row + 1
    if units_row > ws.max_row:
        return units
    for key, col in col_map.items():
        v = ws.cell(row=units_row, column=col).value
        if v is not None:
            units[key] = str(v).strip()
    return units


def parse(filepath: str) -> ParsedSlideSheet:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]

    # --- Header block ---
    raw_header = _scan_header_block(ws, 4, 15)

    # Detect depth unit from labels
    depth_label = raw_header.get("_label_depth_in", raw_header.get("_label_casing_shoe", ""))
    depth_unit = detect_depth_unit(depth_label)
    source_unit_system = "metric" if depth_unit == "m" else "field"

    # --- Find operations table ---
    header_row = _find_ops_header_row(ws)
    if header_row is None:
        raise ValueError(f"Could not find operations table header in {filepath}")

    col_map = _build_col_map(ws, header_row)
    raw_units = _get_units_row(ws, header_row, col_map)

    # Unit detection from ops table units row
    ops_depth_unit = detect_depth_unit(raw_units.get("md_from", raw_units.get("course", "")))
    if ops_depth_unit == "m":
        depth_unit = "m"
        source_unit_system = "metric"

    dls_unit_str = raw_units.get("dls", "deg/100ft")
    dls_unit = detect_dls_unit(dls_unit_str)

    mw_label = raw_header.get("_label_end_mw", "")
    mw_unit = detect_mw_unit(mw_label)

    source_units = {
        "depth": depth_unit,
        "dls": dls_unit,
        "mw": mw_unit,
    }

    # --- Convert header values to canonical ---
    header: dict[str, Any] = {}
    for k, v in raw_header.items():
        if k.startswith("_label_"):
            continue
        label_key = f"_label_{k}"
        label = raw_header.get(label_key, "")
        if k in ("depth_in", "depth_out", "casing_shoe"):
            fv = _safe_float(v)
            header[k] = depth_to_ft(fv, depth_unit) if fv is not None else None
        elif k == "end_mw":
            fv = _safe_float(v)
            header[k + "_ppg"] = mw_to_ppg(fv, mw_unit) if fv is not None else None
        elif k in ("date_in", "date_td", "date_out"):
            header[k] = _parse_date(v)
        elif k in ("bit_run_number",):
            try:
                header[k] = int(v)
            except (TypeError, ValueError):
                header[k] = None
        elif k == "hole_size_in":
            fv = _safe_float(v)
            header[k] = round(fv, 4) if fv is not None else None
        else:
            header[k] = v

    # --- Walk operations rows ---
    data_start = header_row + 2
    surveys: list[SurveyRow] = []
    intervals: list[IntervalRow] = []

    def gcol(key: str) -> Optional[int]:
        return col_map.get(key)

    def gval(row_cells, key: str):
        c = gcol(key)
        if c is None:
            return None
        cell = row_cells[c - 1] if c - 1 < len(row_cells) else None
        return cell.value if cell else None

    # Prepend initial survey row (VBA convention)
    surveys.append(SurveyRow(sequence=0, svy_md_ft=0.0, incl_deg=0.0, azmth_deg=0.0, dls_deg_per_100ft=0.0))

    first_md_from: Optional[float] = None
    survey_seq = 1
    interval_seq = 0
    consecutive_empty = 0

    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row):
        md_from_raw = gval(row, "md_from")
        md_to_raw = gval(row, "md_to")
        svy_md_raw = gval(row, "svy_md")
        mode_raw = gval(row, "operation_mode")

        all_empty = all(c.value is None for c in row)
        if all_empty:
            consecutive_empty += 1
            if consecutive_empty >= 3:
                break
            continue
        consecutive_empty = 0

        # Survey row
        if svy_md_raw is not None:
            md = _safe_float(svy_md_raw)
            inc = _safe_float(gval(row, "incl"))
            azi = _safe_float(gval(row, "azmth"))
            dls = _safe_float(gval(row, "dls"))
            if md is not None:
                # Convert to canonical
                md_ft = depth_to_ft(md, depth_unit)
                dls_100ft = dls_to_100ft(dls or 0.0, dls_unit)
                # Deduplicate: if same MD as last, replace
                if surveys and abs(surveys[-1].svy_md_ft - md_ft) < 0.01:
                    surveys[-1] = SurveyRow(
                        sequence=surveys[-1].sequence,
                        svy_md_ft=md_ft,
                        incl_deg=inc or 0.0,
                        azmth_deg=azi or 0.0,
                        dls_deg_per_100ft=dls_100ft,
                    )
                else:
                    surveys.append(SurveyRow(
                        sequence=survey_seq,
                        svy_md_ft=md_ft,
                        incl_deg=inc or 0.0,
                        azmth_deg=azi or 0.0,
                        dls_deg_per_100ft=dls_100ft,
                    ))
                    survey_seq += 1

        # Interval row
        if md_from_raw is not None and md_to_raw is not None and mode_raw is not None:
            md_from = _safe_float(md_from_raw)
            md_to = _safe_float(md_to_raw)
            if md_from is not None and md_to is not None:
                if first_md_from is None:
                    first_md_from = depth_to_ft(md_from, depth_unit)
                    # Prepend initial rotating interval (VBA convention)
                    intervals.append(IntervalRow(
                        sequence=interval_seq,
                        md_from_ft=0.0,
                        md_to_ft=first_md_from,
                        mode="Rotating",
                    ))
                    interval_seq += 1

                course_raw = gval(row, "course")
                rop_raw = gval(row, "calc_rop")
                tf_mode_raw = gval(row, "tf_mode")
                tf_angle_raw = gval(row, "tf_angle")
                start_raw = gval(row, "start_time")
                end_raw = gval(row, "end_time")

                intervals.append(IntervalRow(
                    sequence=interval_seq,
                    md_from_ft=depth_to_ft(md_from, depth_unit),
                    md_to_ft=depth_to_ft(md_to, depth_unit),
                    mode=_normalize_mode(str(mode_raw)),
                    course_ft=depth_to_ft(_safe_float(course_raw) or 0.0, depth_unit) if course_raw else None,
                    calc_rop=_safe_float(rop_raw),
                    tf_mode=str(tf_mode_raw).strip() if tf_mode_raw else None,
                    tf_angle=_safe_float(tf_angle_raw),
                    start_time=_parse_date(start_raw),
                    end_time=_parse_date(end_raw),
                ))
                interval_seq += 1

    return ParsedSlideSheet(
        header=header,
        surveys=surveys,
        intervals=intervals,
        source_unit_system=source_unit_system,
        source_units=source_units,
        source_filename="",
    )
